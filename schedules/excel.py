import datetime
import os
from celery import current_task
from django.core.cache import cache
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Color, PatternFill, Border, Side, Font
from operator import attrgetter, itemgetter

from .analytics import get_analytics

# TODO: remove file names from these methods

# current_task update helper (used for readability)
def get_percent(current_actions, total_actions):
	return int((current_actions / total_actions) * 100)

# Duplicate the template making a new sheet for each stapher passed in
def create_new_workbook(staphers):
	path = os.path.join(settings.STATIC_ROOT, 'xlsx/schedules-template.xlsx')
	print(f'PATH = {path}')
	file = open(path)
	print(f'FILE = {file}')

	# Setting the initial state to send to the frontend and update the progress bar
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or len(staphers)
	meta = {'message':'Creating New Schedule Workbook', 'process_percent':get_percent(num_actions_made, total_actions)}
	current_task.update_state(meta = meta)

	# Copy the template workbook.
	template_wb = load_workbook('/static/xlsx/schedules-template.xlsx')
	template_wb.save("/static/xlsx/schedules.xlsx")
	schedule_wb = load_workbook('/static/xlsx/schedules.xlsx')
	template_ws = schedule_wb['TEMPLATE']

	for i, stapher in enumerate(staphers):
		# Update the state of progress for the front end
		message = f'Creating Excel Worksheet for {stapher.full_name()}'
		meta = {'message':message, 'process_percent':get_percent(num_actions_made + i, total_actions)}
		current_task.update_state(meta = meta)
		
		# We copy the template worksheet for each Stapher.
		stapher_ws = schedule_wb.copy_worksheet(template_ws)
		stapher_ws.title = stapher.full_name()
	schedule_wb.remove(template_ws)

	# Updating the state to send to the frontend and update the progress bar
	meta = {'message':'Saving New Schedule Workbook', 'process_percent':get_percent(len(staphers), total_actions)}
	current_task.update_state(meta = meta)
	cache.set('num_actions_made', num_actions_made + len(staphers), None)

	# Save the workbook and return it's destination
	schedule_wb.save("/static/xlsx/schedules.xlsx")
	return '/static/xlsx/schedules.xlsx'

# update_individual_excel_files helper
def get_row_from_day(day):
	return 5 + (day * 4)

# update_individual_excel_files helper
def get_start_col_from_time(time):
	return 3 + ((time.hour - 6) * 4) + int((time.minute / 60) * 4)

# update_individual_excel_files helper
def get_end_col_from_time(time):
	return get_start_col_from_time(time) - 1

# This function takes in a list of staphers and staphings and makes a readable xl file for each stapher.
def update_individual_excel_files(staphers, staphings):
	# Copy the template workbook
	wb_str = create_new_workbook(staphers)

	# Setting the initial state to send to the frontend and update the progress bar
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or len(staphers)
	meta = {'message':'Loading Schedule Workbook', 'process_percent':get_percent(num_actions_made, total_actions)}
	current_task.update_state(meta = meta)

	schedule_wb = load_workbook("/static/xlsx/schedules.xlsx")
	seconds_in_hour = 60 * 60
	for i, stapher in enumerate(staphers):
		# Update the state of progress for the front end
		message = f'Populating Excel Worksheet for {stapher.full_name()}'
		meta = {'message':message, 'process_percent':get_percent(num_actions_made + i, total_actions)}
		current_task.update_state(meta = meta)
		
		# Load the worksheet for the stapher
		stapher_ws = schedule_wb[stapher.full_name()]
		stapher_ws['A1'] = 'Name: ' + stapher.full_name()
		stapher_ws['AE1'] = 'Postion: ' + stapher.title

		# Edit the excel sheet for each scheduled shift
		shifts = stapher.ordered_shifts(staphings)
		for shift in shifts:
			row = get_row_from_day(shift.day)
			start_col = get_start_col_from_time(shift.start)
			end_col = get_end_col_from_time(shift.end)
			cell = stapher_ws.cell(row = row, column = start_col)
			cell.alignment = Alignment(shrinkToFit = True, wrapText = True, horizontal = 'center', vertical = 'center')
			cell.border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'), bottom = Side(style = 'thin'))

			# Resize the text in the cell to make it more readable.
			font_size = 10
			value = shift.get_excel_str()
			if shift.length().seconds <= seconds_in_hour:
				font_size = 8
			if shift.length().seconds <= seconds_in_hour / 2:
				font_size = 6
				value = shift.title
			cell.font = Font(size=font_size)
			cell.value = value

			# Next we make changes for unpaid shifts and add a 1 to the appropriate time cell
			if shift.is_unpaid():
				cell.fill = PatternFill(patternType = 'solid', fill_type = 'solid', fgColor = Color('C4C4C4'))
			else:
				if shift.is_programming():
					row_adjust = 2
				else:
					row_adjust = 1
				for col in range(start_col, end_col + 1):
					time_cell = stapher_ws.cell(row = row - row_adjust, column = col)
					time_cell.value = '1'

			stapher_ws.merge_cells(start_row = row, start_column = start_col, end_row = row, end_column = end_col)
	

	# Updating the state to send to the frontend and update the progress bar
	meta = {'message':'Saving New Schedule Workbook', 'process_percent':get_percent(num_actions_made + len(staphers), total_actions)}
	current_task.update_state(meta = meta)
	cache.set('num_actions_made', num_actions_made + len(staphers), None)

	# Save the workbook
	schedule_wb.save("/static/xlsx/schedules.xlsx")

# This function takes in a set of staphings and returns a list of the staphers working them 
def get_shift_workers(staphings):
	shifts_and_workers = {}
	for staphing in staphings:
		if staphing.shift.id in shifts_and_workers:
			shifts_and_workers[staphing.shift.id].append(staphing.stapher)
		else:
			shifts_and_workers[staphing.shift.id] = [staphing.stapher]
	return shifts_and_workers

# This returns a dictionary of shift ids to the shifts that are at that time
def get_shifts_ids_to_height(ordered, shifts):
	section_height = 10
	ids_to_height = {}
	largest_overlap = 0
	for day in ordered:
		times_for_day = ordered[day]
		for i in range(1, len(times_for_day)):
			start = times_for_day[i - 1]
			end = times_for_day[i]
			shifts_during_window = [shift for shift in shifts if shift.is_in_window(day, start, end)]
			if largest_overlap < len(shifts_during_window):
				largest_overlap = len(shifts_during_window)
			for shift in shifts_during_window:
				height = int(section_height / len(shifts_during_window))
				if shift.id in ids_to_height:
					old_height = ids_to_height[shift.id]
					if old_height > height:
						ids_to_height[shift.id] = height
				else:
					ids_to_height[shift.id] = height
	return ids_to_height

def set_times_to_offsets(ordered):
	times_to_offsets = {}
	for day in ordered:
		times_for_day = ordered[day]
		for i in range(1, len(times_for_day)):
			start = times_for_day[i - 1]
			time_key = get_seconds_from_day_and_time(day, start)
			times_to_offsets[time_key] = 0
	return times_to_offsets

def get_all_time_keys(shift, times_to_offset):
	time_keys = []
	shift_start_key = get_seconds_from_day_and_time(shift.day, shift.start)
	shift_end_key = get_seconds_from_day_and_time(shift.day, shift.end)
	for time_key in times_to_offset:
		if time_key >= shift_start_key and time_key < shift_end_key:
			time_keys.append(time_key)
	return time_keys

def get_and_update_largest_offset(shift, times_to_offset, height):
	time_keys = get_all_time_keys(shift, times_to_offset)
	largest_offset = 0
	for key in time_keys:
		if times_to_offset[key] > largest_offset:
			largest_offset = times_to_offset[key]
	for key in time_keys:
		# We only want to update the cells that are on the same level
		if times_to_offset[key] == largest_offset:
			times_to_offset[key] = times_to_offset[key] + height
	return largest_offset

# TODO: DRY with create_new_workbook method
def copy_master_template(masters):
	# Set the progress for the frontend 
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or len(masters)
	meta = {'message':'Copying Master Template', 'process_percent':get_percent(num_actions_made, total_actions)}
	current_task.update_state(meta = meta)

	# Copy the template workbook.
	template_wb = load_workbook('/static/xlsx/masters-template.xlsx')
	template_wb.save("/static/xlsx/masters.xlsx")
	master_wb = load_workbook('/static/xlsx/masters.xlsx')
	for i, master in enumerate(masters):
		# Update the progress for each master
		num_actions_made = cache.get('num_actions_made') or 0
		total_actions = cache.get('num_total_actions') or len(masters)
		message = f'Creating Master Worksheet for {master}'
		meta = {'message':message, 'process_percent':get_percent(num_actions_made + i, total_actions)}
		current_task.update_state(meta = meta)

		# We copy the template worksheet for each master.
		template_ws = master_wb['TEMPLATE']	
		master_ws = master_wb.copy_worksheet(template_ws)
		master_ws.title = master.title

	# Update actions made in the cache
	meta = {'message':'Saving New Master Workbook', 'process_percent':get_percent(num_actions_made + len(masters), total_actions)}
	current_task.update_state(meta = meta)
	cache.set('num_actions_made', num_actions_made + len(masters), None)

	# Save the workbook
	master_wb.remove(template_ws)
	master_wb.save("/static/xlsx/masters.xlsx")

# TODO: DRY with get_start_col_from_time method
def get_master_start_col_from_time(time):
	return 2 + ((time.hour - 6) * 4) + int((time.minute / 60) * 4)

def get_master_end_col_from_time(time):
	return get_master_start_col_from_time(time) - 1

def master_row(shift):
	return (shift.day * 13) + 5

def get_col_range(shift):
	start_col = get_master_start_col_from_time(shift.start)
	end_col = get_master_end_col_from_time(shift.end)
	return [col for col in range(start_col, end_col + 1)]

# This function returns a dictionary of shift id to the height and offset of the shift
def get_shifts_ids_to_placement(shifts):
	section_height = 12
	cols_to_info = {}
	for shift in shifts:
		col_range = get_col_range(shift)
		if shift.day not in cols_to_info:
			cols_to_info[shift.day] = {}
		for col in col_range:
			if col not in cols_to_info[shift.day]:
				# The first place in the list represents the number of shifts at this time, the second represents the taken rows
				cols_to_info[shift.day][col] = [0, []] 
			cols_to_info[shift.day][col][0] = cols_to_info[shift.day][col][0] + 1
	ids_to_placement = {}
	shifts = sorted(shifts, key = attrgetter('day', 'start'))
	# In order to avoid overflowing the excel sheet - we first place the top layer of each day and then the 2nd layer and so on...
	placed_shifts = []
	last_shift = shifts[0]
	while len(placed_shifts) < len(shifts):
		for shift in shifts:
			not_seen_before = shift not in placed_shifts
			should_place_next  = shift.start == last_shift.start or not shift.overlaps(last_shift)
			have_to_place = len(shifts) == len(placed_shifts) + 1 # Since it is the last shift we have to place it regardless
			if not_seen_before and (should_place_next or have_to_place):
				# First we get the height...
				col_range = get_col_range(shift)
				most_overlap_in_range = 1
				for col in col_range:
					overlap_at_col = cols_to_info[shift.day][col][0]
					if most_overlap_in_range < overlap_at_col:
						most_overlap_in_range = overlap_at_col
				shift_height = int(section_height / most_overlap_in_range)
				# Now we find the offset by checking for the first availible set of rows 
				offset = 0
				solution_found = False
				while not solution_found:
					rows_to_occupy = [row for row in range(offset, offset + shift_height)]
					overlap_found = False
					for col in col_range:
						occupied_rows = cols_to_info[shift.day][col][1]
						if bool(set(rows_to_occupy) & set(occupied_rows)):
							overlap_found = True
					if not overlap_found:
						solution_found = True
					else:
						offset += 1
				for col in col_range:
					cols_to_info[shift.day][col][1].extend(rows_to_occupy)

				# Now we find the start row/col and end row/col and store that information to the shifts ID
				start_col = col_range[0]
				end_col = col_range[-1]
				start_row = master_row(shift) + offset
				end_row = start_row + shift_height - 1
				ids_to_placement[shift.id] = [start_col, start_row, end_col, end_row]

				# If the shifts have the same start time, we don't need to update the shift unless the one we're looking at ends earlier
				if shift.start != last_shift.start or shift.end < last_shift.end:
					last_shift = shift
				placed_shifts.append(shift)
	return ids_to_placement

def get_length(shift):
	return shift.length()

def update_standard_masters(masters, staphings):
	# Copy the master template
	masters =  sorted(masters, key=attrgetter('title'))
	copy_master_template(masters)
	

	# Set the ammount of actions taken / needed to be take to send to the front end
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or len(masters)
	meta = {'message':'Loading New Master Workbook', 'process_percent':get_percent(num_actions_made, total_actions)}
	current_task.update_state(meta = meta)

	# Load the new master workbook
	master_wb = load_workbook('/static/xlsx/masters.xlsx')

	# Update the masters
	for i, master in enumerate(masters):
		message = f'Updating Master Worksheet for {master}'
		meta = {'message':message, 'process_percent':get_percent(num_actions_made + i,total_actions)}
		current_task.update_state(meta = meta)
		master_staphings = master.get_master_staphings(staphings)
		shift_workers = get_shift_workers(master_staphings)
		master_shifts = list(set([s.shift for s in master_staphings]))
		master_shifts = [shift[0] for shift in sorted([[shift, shift.start, shift.length()] for shift in master_shifts], key = itemgetter(1, 2))]
		ids_to_placement = get_shifts_ids_to_placement(master_shifts)
		master_ws = master_wb[master.title]
		master_ws['A1'] =  master.title + ' Master'
		for shift in master_shifts:
			start_col = ids_to_placement[shift.id][0]
			start_row = ids_to_placement[shift.id][1]
			end_col = ids_to_placement[shift.id][2]
			end_row = ids_to_placement[shift.id][3]
			cell = master_ws.cell(row = start_row, column = start_col)
			value = shift.title + ':\n'

			# If the shift is longer than an hour... 
			if start_col < end_col - 3:
				value = shift.get_excel_str() + ':\n'

			# If the shift is 30 min or shorter... 
			if start_col >= end_col - 1:
				cell.font = Font(size = 7)
			else:
				cell.font = Font(size = 12)

			# Now we add the names of the people working the shift
			for worker in set(shift_workers[shift.id]):
				value = value + worker.first_name + ', '
			value = value[:-2]
			cell.value = value
			cell.alignment = Alignment(shrinkToFit = True, wrapText = True, horizontal = 'center', vertical = 'center')
			cell.border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'), bottom = Side(style = 'thin'))
			master_ws.merge_cells(start_row = start_row, start_column = start_col, end_row = end_row, end_column = end_col)
	
	# Reset the cache and send the final message to the front end
	cache.set('num_actions_made', num_actions_made + len(masters))
	meta = {'message':'Saving New Master Workbook', 'process_percent':get_percent(num_actions_made + len(masters), total_actions)}
	current_task.update_state(meta = meta)

	# Save the new master workbook
	master_wb.save("/static/xlsx/masters.xlsx")
			
def get_meal_master_col(shift):
	return shift.day + 2

def get_meal_master_starting_row(shift):
	if shift.has_flag('meal-head'):
		return 3
	elif shift.has_flag('hobart'):
		if shift.start == datetime.time(12, 0, 0, 0):
			return 25
		elif shift.start == datetime.time(17, 30, 0, 0):
			return 24
	elif shift.has_flag('wine'):
			return 22
	if shift.start == datetime.time(6, 45, 0, 0):
		return 5
	elif shift.start == datetime.time(7, 0, 0, 0):
		return 20
	elif shift.start == datetime.time(7, 30, 0, 0):
		return 8
	elif shift.start == datetime.time(7, 45, 0, 0):
		return 12
	elif shift.start == datetime.time(8, 15, 0, 0):
		return 16
	elif shift.start == datetime.time(11, 15, 0, 0):
		return 5
	elif shift.start == datetime.time(11, 45, 0, 0):
		return 10
	elif shift.start == datetime.time(12, 0, 0, 0):
		return 19
	elif shift.start == datetime.time(17, 15, 0, 0):
		return 4
	elif shift.start == datetime.time(17, 30, 0, 0):
		return 13
	elif shift.start == datetime.time(18, 0, 0, 0):
		return 15


def update_meal_masters(masters, staphings):
	# Set the ammount of actions taken / needed to be take to send to the front end
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or len(masters)
	meta = {'message':'Creating New Master Workbook', 'process_percent':get_percent(num_actions_made, total_actions)}
	current_task.update_state(meta = meta)
	
	# Load the meal master workbook
	meal_master_wb = load_workbook('/static/xlsx/meal-master-template.xlsx')

	# Update the workbook
	for i, master in enumerate(masters):
		# Update the progress and send to the front end
		message = f'Updating {master} Master'
		meta = {'message':message, 'process_percent':get_percent(num_actions_made + i, total_actions)}
		current_task.update_state(meta = meta)

		# Get the correct worksheet for each master 
		master_ws = meal_master_wb[master.title]
		master_staphings = master.get_master_staphings(staphings)
		shift_workers = get_shift_workers(master_staphings)
		master_shifts = list(set([s.shift for s in master_staphings]))

		# Place each shift in its proper group of cells
		for shift in master_shifts:
			col = get_meal_master_col(shift)
			curr_row = get_meal_master_starting_row(shift)
			workers = set(shift_workers[shift.id])

			# Add the worker's names to the cell
			for worker in workers:
				cell = master_ws.cell(row = curr_row, column = col)
				cell.value = worker.full_name()
				cell.font = Font(size = 12)
				curr_row += 1

	# Reset the cache and send the final message to the front end
	cache.set('num_actions_made', num_actions_made + len(masters))
	meta = {'message':'Saving New Meal Master Workbook', 'process_percent':get_percent(num_actions_made + len(masters), total_actions)}
	current_task.update_state(meta = meta)

	# Save the meal master workbook
	meal_master_wb.save("/static/xlsx/meal-masters.xlsx")


def update_masters(masters, staphings):
	standard_masters = []
	meal_masters = []
	for master in masters:
		if master.is_standard():
			standard_masters.append(master)
		else:
			meal_masters.append(master)
	update_standard_masters(standard_masters, staphings)
	update_meal_masters(meal_masters, staphings)

# Get the analytics for the given schedule and place them into an excel spreadsheet
def update_analytics(staphers, staphings, flags, qualifications):
	# Set the values and update the progress for the front end
	num_actions_made = cache.get('num_actions_made') or 0
	total_actions = cache.get('num_total_actions') or 2
	meta = {'message':'Retrieving Analytics', 'process_percent':get_percent(num_actions_made + 1, total_actions)}
	current_task.update_state(meta = meta)

	# Retrieve the analytics
	analytics = get_analytics(staphers, staphings, flags, qualifications)

	#Load/Update the workbook
	analytics_wb = load_workbook('/static/xlsx/analytics-template.xlsx')
	analytics_ws = analytics_wb['Analytics']
	analytics_ws.title = 'Analytics'

	# Update the progress of the task and send it to the front end
	meta = {'message':'Updating Analytics Workbook', 'process_percent':get_percent(num_actions_made + 2, total_actions)}
	current_task.update_state(meta = meta)

	# For each analytics, update the worksheet
	for row in range(0, len(analytics)):
		
		for col in range(0, len(analytics[row])):
			cell = analytics_ws.cell(row = row + 2, column = col + 1)
			cell.value = analytics[row][col]
			if row == 0:
				cell.font = Font(size = 12)
				cell.alignment = Alignment(shrinkToFit = True, wrapText = True, horizontal = 'center', vertical = 'top')
				cell.border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'), bottom = Side(style = 'thin'))

	# Update the front end progress and reset the cached number of actions taken
	meta = {'message':'Saving Analytics Workbook', 'process_percent':get_percent(num_actions_made + 3, total_actions)}
	current_task.update_state(meta = meta)
	cache.set('num_actions_made', num_actions_made + 3, None)

	# Save the workbook
	analytics_wb.save("/static/xlsx/analytics.xlsx")





